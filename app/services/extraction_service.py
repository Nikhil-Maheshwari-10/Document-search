import csv
import PyPDF2
from openpyxl import load_workbook
from langchain_text_splitters import RecursiveCharacterTextSplitter
from app.config import CHUNK_SIZE, CHUNK_OVERLAP
from app.logger import logger

def extract_text_from_file(file_path):
    """Extract text content from a file based on its extension"""
    try:
        file_extension = file_path.rsplit('.', 1)[1].lower()
        
        if file_extension == 'txt':
            try:
                with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
                    return file.read()
            except Exception as e:
                logger.error(f"Error reading txt file: {e}")
                return ""
                
        elif file_extension == 'pdf':
            try:
                text = ""
                with open(file_path, 'rb') as file: 
                    pdf_reader = PyPDF2.PdfReader(file)
                    for page_num in range(len(pdf_reader.pages)):
                        text += pdf_reader.pages[page_num].extract_text() + "\n"
                return text
            except Exception as e:
                logger.error(f"Error reading PDF file: {e}")
                return ""
            
        elif file_extension == 'csv':
            try:
                text = ""
                with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
                    reader = csv.reader(file)
                    for row in reader:
                        text += ", ".join(row) + "\n"
                return text
            except Exception as e:
                logger.error(f"Error reading CSV file: {e}")
                return ""
            
        elif file_extension == 'xlsx':
            try:
                text = ""
                workbook = load_workbook(file_path, read_only=True)
                for sheet in workbook:
                    for row in sheet.iter_rows(values_only=True):
                        text += ", ".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
                return text
            except Exception as e:
                logger.error(f"Error reading XLSX file: {e}")
                return ""
        
        return ""
    except Exception as e:
        logger.error(f"Error extracting text from file: {e}")
        return ""

def create_chunks(text, chunk_size=CHUNK_SIZE, overlap=CHUNK_OVERLAP):
    """Split text into chunks with overlap using LangChain's RecursiveCharacterTextSplitter"""
    if not text:
        return []
    
    try:    
        text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=overlap,
            length_function=len,
        )
        chunks = text_splitter.split_text(text)
        return chunks
    except Exception as e:
        logger.error(f"Error in create_chunks: {e}")
        # Fallback to a simple chunking method
        chunks = []
        start = 0
        text_length = len(text)
        while start < text_length:
            end = min(start + chunk_size, text_length)
            chunks.append(text[start:end])
            start = end - overlap if end < text_length else text_length
        return chunks
